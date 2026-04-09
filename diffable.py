"""
diffable — Generate interactive HTML diffs from versioned JSON data.

Usage:
    from diffable import DiffTable
    DiffTable("eGPIO.json", title="eGPIO Register Map", key="index").generate()

JSON format:
    {
      "versions": [
        {
          "version": "v1.0",
          "date": "June 2025",
          "<data_key>": [
            {"index": 0, "signal": "SIG_A", "note": "some note", ...}
          ]
        },
        ...
      ]
    }

The library auto-detects the data array key (any key that is not "version" or "date").
The "note" field, if present, is shown in the right-side detail panel.
"""

import csv
import json
from html import escape as _html_escape
from pathlib import Path

from _templates import STYLE, JS_TEMPLATE, HTML_TEMPLATE

_META_KEYS = {"version", "date"}


def _esc(s):
    """HTML-escape a string, returning '' for None."""
    return _html_escape(str(s), quote=True) if s is not None else ""


def _safe_name(name):
    """Sanitize a string for use as a filename."""
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name)


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for xlsx support. "
            "Install it with: pip install openpyxl"
        )


def _read_sheet_rows(ws):
    """Read an openpyxl worksheet into a list of dicts."""
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    headers = [str(h) if h is not None else f"col_{i}"
               for i, h in enumerate(headers)]

    raw_rows = []
    for row in rows_iter:
        record = {}
        for h, val in zip(headers, row):
            if val is None:
                record[h] = ""
            else:
                record[h] = val if isinstance(val, (int, float, bool)) else str(val)
        raw_rows.append(record)

    # Remove empty columns (all values are empty after stripping)
    non_empty_cols = [h for h in headers
                      if any(str(r.get(h, "")).strip() for r in raw_rows)]
    # Remove empty rows (all values are empty after stripping)
    rows = []
    for r in raw_rows:
        if any(str(r.get(h, "")).strip() for h in non_empty_cols):
            rows.append({h: r[h] for h in non_empty_cols})

    return rows


def _make_version_entry(version, date, data_key, rows):
    """Build a version entry dict."""
    entry = {"version": version}
    if date:
        entry["date"] = date
    entry[data_key] = rows
    return entry


def _detect_data_key(version_obj):
    """Find the first key in a version object that holds the data array."""
    for k, v in version_obj.items():
        if k not in _META_KEYS and isinstance(v, list):
            return k
    raise ValueError(
        "Could not detect data key. Each version must contain an array field "
        "besides 'version' and 'date'."
    )


def _detect_columns(versions, data_key):
    """Collect all field names across every row in every version, preserving order."""
    seen = {}
    order = 0
    for ver in versions:
        for row in ver.get(data_key, []):
            for k in row:
                if k not in seen:
                    seen[k] = order
                    order += 1
    return [k for k, _ in sorted(seen.items(), key=lambda x: x[1])]


class DiffTable:
    """Generate an interactive, version-diffing HTML page from a JSON file.

    Parameters
    ----------
    json_source : str or Path — path to the JSON file.
    title : str — page/header title.
    key : str — field used to match rows across versions. Defaults to the first field.
    output : str or Path or None — output HTML path. Defaults to ``<title>.html``.
    columns : list[str] or None — explicit columns to display. Auto-detected if None.
    note_field : str — field shown in the side panel. Default "note".
    """

    def __init__(self, json_source, *, title="Diff Explorer", key=None,
                 output=None, columns=None, note_field="note"):
        self.json_source = Path(json_source)
        self.title = title
        self._key = key
        self._output = Path(output) if output else None
        self._columns = columns
        self.note_field = note_field

    def generate(self):
        with self.json_source.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if "versions" not in data or not data["versions"]:
            raise ValueError("JSON must contain a non-empty 'versions' array.")

        versions = data["versions"]
        data_key = _detect_data_key(versions[0])
        all_cols = _detect_columns(versions, data_key)

        key = self._key or all_cols[0]
        if key not in all_cols:
            raise ValueError(f"Key field '{key}' not found in data.")

        if self._columns:
            display_cols = [c for c in self._columns if c in all_cols]
        else:
            display_cols = [c for c in all_cols if c != self.note_field]

        if key in display_cols:
            display_cols.remove(key)
        display_cols.insert(0, key)

        out_path = self._output or (self.json_source.parent / f"{self.title}.html")
        html = self._render(versions, data_key, key, display_cols)
        out_path.write_text(html, encoding="utf-8")
        return out_path

    def _render(self, versions, data_key, key, display_cols):
        col_headers = "".join(
            f"<th>{self._col_label(c)}</th>" for c in display_cols
        )
        col_labels = [c.replace("_", " ").title() for c in display_cols]
        script = JS_TEMPLATE.substitute(
            KEY=json.dumps(key),
            DATA_KEY=json.dumps(data_key),
            COLS=json.dumps(display_cols),
            COL_LABELS=json.dumps(col_labels),
            NOTE_FIELD=json.dumps(self.note_field),
            VERSIONS=json.dumps(versions),
        )
        return HTML_TEMPLATE.substitute(
            TITLE=_esc(self.title),
            STYLE=STYLE,
            COL_HEADERS=col_headers,
            SCRIPT=script,
        )

    @staticmethod
    def _col_label(name):
        return _esc(name.replace("_", " ").title())


class SpreadsheetConverter:
    """Convert xlsx/csv files to the versioned JSON format used by DiffTable.

    Parameters
    ----------
    source : str or Path — path to an .xlsx, .xls, or .csv file.
    output_dir : str or Path or None — defaults to the source file's directory.
    version : str — version label (e.g. "v1.0").
    date : str or None — optional date string.
    data_key : str — key name for the data array. Default "data".
    """

    def __init__(self, source, *, output_dir=None, version="v1.0", date=None,
                 data_key="data"):
        self.source = Path(source)
        self.output_dir = Path(output_dir) if output_dir else self.source.parent
        self.version = version
        self.date = date
        self.data_key = data_key

    def convert(self, append=False):
        """Convert the source file and write JSON files. Returns list[Path]."""
        suffix = self.source.suffix.lower()
        if suffix == ".csv":
            sheets = {"": self._read_csv()}
        elif suffix in (".xlsx", ".xls"):
            sheets = self._read_excel()
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

        stem = self.source.stem
        outputs = []
        for sheet_name, rows in sheets.items():
            filename = f"{stem}_{_safe_name(sheet_name)}.json" if sheet_name else f"{stem}.json"
            out_path = self.output_dir / filename

            version_entry = _make_version_entry(self.version, self.date, self.data_key, rows)

            if append and out_path.exists():
                with out_path.open("r", encoding="utf-8") as f:
                    existing = json.load(f)
                existing["versions"].append(version_entry)
            else:
                existing = {"versions": [version_entry]}

            with out_path.open("w", encoding="utf-8") as f:
                json.dump(existing, f, indent=2, ensure_ascii=False)

            outputs.append(out_path)

        return outputs

    def _read_csv(self):
        with self.source.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader]

    def _read_excel(self):
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(self.source, read_only=True, data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            rows = _read_sheet_rows(wb[name])
            if rows:
                sheets[name] = rows
        wb.close()
        return sheets


class ExcelDiff:
    """Scan a directory of xlsx files -> JSON -> HTML diffs.

    Files are sorted by name to determine version order. Sheets with the same
    name across files become successive versions.

    Parameters
    ----------
    directory : str or Path — directory containing .xlsx files.
    output_dir : str or Path or None — defaults to ``directory``.
    key : str or None — row identity field. Defaults to the first column.
    data_key : str — key for the data array. Default "data".
    note_field : str — field shown in the side panel. Default "note".
    version_func : callable or None — ``(filepath) -> (version_label, date | None)``.
    recent : int or None — only use the last N files (by name). None means all.
    """

    def __init__(self, directory, *, output_dir=None, key=None, data_key="data",
                 note_field="note", version_func=None, recent=None):
        self.directory = Path(directory)
        self.output_dir = Path(output_dir) if output_dir else self.directory
        self.key = key
        self.data_key = data_key
        self.note_field = note_field
        self.version_func = version_func or (lambda fp: (fp.stem, None))
        self.recent = recent

    def run(self):
        """Execute the full pipeline: xlsx -> JSON -> HTML. Returns list[Path]."""
        openpyxl = _require_openpyxl()

        files = sorted(
            (f for f in self.directory.iterdir()
             if f.is_file() and f.suffix.lower() == ".xlsx"),
            key=lambda f: f.name,
        )
        if not files:
            raise FileNotFoundError(f"No .xlsx files found in {self.directory}")

        if self.recent is not None:
            files = files[-self.recent:]

        self.output_dir.mkdir(parents=True, exist_ok=True)

        # sheet_name -> {"versions": [...]}
        sheet_data = {}

        for filepath in files:
            version_label, date_str = self.version_func(filepath)
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                rows = _read_sheet_rows(wb[sheet_name])
                if not rows:
                    continue

                version_entry = _make_version_entry(
                    version_label, date_str, self.data_key, rows
                )

                safe = _safe_name(sheet_name)
                if safe not in sheet_data:
                    sheet_data[safe] = {"name": sheet_name, "versions": []}
                sheet_data[safe]["versions"].append(version_entry)

            wb.close()

        html_outputs = []
        for safe_name, info in sheet_data.items():
            json_path = self.output_dir / f"{safe_name}.json"
            with json_path.open("w", encoding="utf-8") as f:
                json.dump({"versions": info["versions"]}, f, indent=2,
                          ensure_ascii=False)

            dt = DiffTable(
                json_path,
                title=info["name"],
                key=self.key,
                note_field=self.note_field,
            )
            html_outputs.append(dt.generate())

        return html_outputs


if __name__ == "__main__":
    filepath = Path(__file__).parent / "eGPIO.json"
    t = DiffTable(filepath, title="eGPIO Register Map", key="index")
    out = t.generate()

    filepath = Path(__file__).parent / "eRST.json"
    t = DiffTable(filepath, title="eRST Register Map", key="index")
    out = t.generate()
