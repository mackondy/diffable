"""
diffable — Generate interactive HTML diffs from versioned JSON data.

Usage:
    from diffable import DiffTable
    DiffTable("eGPIO.json", title="eGPIO Register Map", key="index").generate()

JSON format:
    {
      "versions": [
        {
          "version": "v1.0",
          "date": "June 2025",
          "<data_key>": [
            {"index": 0, "signal": "SIG_A", "note": "some note", ...}
          ]
        },
        ...
      ]
    }

The library auto-detects the data array key (any key that is not "version" or "date").
The "note" field, if present, is shown in the right-side detail panel.
"""

import csv
import json
import re
import shutil
import subprocess
import zipfile
from collections import Counter
from html import escape as _html_escape
from pathlib import Path

from ._templates import STYLE, JS_TEMPLATE, HTML_TEMPLATE, TOGGLE_PILL_HTML

_META_KEYS = {"version", "date"}

_GITIGNORE_SRC = Path(__file__).resolve().parent.parent / "docs" / "gitignore.txt"


def _esc(s):
    """HTML-escape a string, returning '' for None."""
    return _html_escape(str(s), quote=True) if s is not None else ""


def _safe_name(name):
    """Sanitize a string for use as a filename."""
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name)


def _is_visible(zip_path):
    """True if no path component is hidden (dot-prefixed) or a macOS metadata dir."""
    parts = Path(zip_path).parts
    return not any(p.startswith(".") or p == "__MACOSX" for p in parts)


def _detect_wrapper(visible):
    """If every visible member shares one top-level directory, return its name.

    Returns ``None`` otherwise (multiple tops, or the sole top is a single file).
    """
    if not visible:
        return None
    tops = {Path(n).parts[0] for n in visible if Path(n).parts}
    if len(tops) != 1:
        return None
    wrapper = next(iter(tops))
    # Confirm the wrapper has children — otherwise it's a single top-level file.
    has_children = any(
        len(Path(n).parts) > 1 and Path(n).parts[0] == wrapper
        for n in visible
    )
    return wrapper if has_children else None


def _top_level_entries(visible, wrapper):
    """Top-level entry names in the zip, relative to the wrapper if present."""
    result = set()
    for n in visible:
        parts = Path(n).parts
        if not parts:
            continue
        if wrapper:
            if parts[0] != wrapper or len(parts) < 2:
                continue
            result.add(parts[1])
        else:
            result.add(parts[0])
    return result


_VERSION_RE = re.compile(r"v\d+(?:\.\d+)+")


def _normalize_version(tag):
    """Strip leading zeros from each numeric component (``v1.02`` -> ``v1.2``)."""
    first, *rest = tag.split(".")
    first = "v" + str(int(first[1:])) if first.startswith("v") else first
    return ".".join([first, *(str(int(p)) for p in rest)])


def _default_branch_func(filepath):
    """Extract a normalized version tag from the filename stem.

    Splits the stem on ``_`` and returns the first token that strictly matches
    a version tag like ``v1.00`` or ``v1.0.2``, with leading zeros stripped
    from each component (``v1.02`` -> ``v1.2``, ``v1.02.0`` -> ``v1.2.0``).
    Falls back to the full stem if no version token is found.
    """
    for token in filepath.stem.split("_"):
        if _VERSION_RE.fullmatch(token):
            return _normalize_version(token)
    return filepath.stem


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for xlsx support. "
            "Install it with: pip install openpyxl"
        )


def _read_sheet_rows(ws):
    """Read an openpyxl worksheet into a list of dicts."""
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    headers = [str(h) if h is not None else f"col_{i}"
               for i, h in enumerate(headers)]

    raw_rows = []
    for row in rows_iter:
        record = {}
        for h, val in zip(headers, row):
            if val is None:
                record[h] = ""
            else:
                record[h] = val if isinstance(val, (int, float, bool)) else str(val)
        raw_rows.append(record)

    # Remove empty columns (all values are empty after stripping)
    non_empty_cols = [h for h in headers
                      if any(str(r.get(h, "")).strip() for r in raw_rows)]
    # Remove empty rows (all values are empty after stripping)
    rows = []
    for r in raw_rows:
        if any(str(r.get(h, "")).strip() for h in non_empty_cols):
            rows.append({h: r[h] for h in non_empty_cols})

    return rows


def _make_version_entry(version, date, data_key, rows):
    """Build a version entry dict."""
    entry = {"version": version}
    if date:
        entry["date"] = date
    entry[data_key] = rows
    return entry


def _detect_data_key(version_obj):
    """Find the first key in a version object that holds the data array."""
    for k, v in version_obj.items():
        if k not in _META_KEYS and isinstance(v, list):
            return k
    raise ValueError(
        "Could not detect data key. Each version must contain an array field "
        "besides 'version' and 'date'."
    )


def _detect_columns(versions, data_key):
    """Collect all field names across every row in every version, preserving order."""
    seen = {}
    order = 0
    for ver in versions:
        for row in ver.get(data_key, []):
            for k in row:
                if k not in seen:
                    seen[k] = order
                    order += 1
    return [k for k, _ in sorted(seen.items(), key=lambda x: x[1])]


def _norm(v):
    """Normalize a cell value for change detection — match the JS ``String(x ?? '')``."""
    return "" if v is None else str(v)


def _build_diff_only_versions(versions, data_key, key, note_field="note"):
    """Strip each version down to its changed rows vs the previous version.

    Returns a list of versions where each entry has a ``diff_rows`` list
    instead of the full data array. Each diff entry is one of::

        {"status": "added",    "curr": {row}}
        {"status": "modified", "prev": {row}, "curr": {row}}
        {"status": "removed",  "prev": {row}}

    The first version additionally carries a filtered data array under
    ``data_key`` containing only rows whose keys appear in some diff
    entry across the timeline — i.e. the baseline state of every record
    that ever changes. Records identical across all versions are dropped
    entirely, which is the whole point of changes_only mode. Spacer rows
    are dropped too.

    ``note_field`` is excluded from the modification check — if two rows
    differ only in their note (auxiliary metadata, not displayed as a
    column), they are treated as identical. Useful for e.g. resistor
    refdes mappings that vary across product but represent the same
    component role.
    """
    new_versions = []
    prev_rows_by_key = {}

    for vi, ver in enumerate(versions):
        rows = ver.get(data_key, [])

        curr_rows_by_key = {}
        order_keys = []
        occ = {}
        for r in rows:
            k = r.get(key)
            if k is None or k == "":
                continue
            occ[k] = occ.get(k, 0) + 1
            ck = (k, occ[k])
            curr_rows_by_key[ck] = r
            order_keys.append(ck)

        diff_rows = []
        if vi > 0:
            for ck in order_keys:
                curr = curr_rows_by_key[ck]
                prev = prev_rows_by_key.get(ck)
                if prev is None:
                    diff_rows.append({"status": "added", "curr": curr})
                else:
                    cols = (set(prev) | set(curr)) - {note_field}
                    if any(_norm(prev.get(c)) != _norm(curr.get(c)) for c in cols):
                        diff_rows.append({
                            "status": "modified",
                            "prev": prev,
                            "curr": curr,
                        })

            for pk, prev in prev_rows_by_key.items():
                if pk not in curr_rows_by_key:
                    diff_rows.append({"status": "removed", "prev": prev})

        new_ver = {"version": ver["version"]}
        if "date" in ver:
            new_ver["date"] = ver["date"]
        new_ver["diff_rows"] = diff_rows
        new_versions.append(new_ver)

        prev_rows_by_key = curr_rows_by_key

    # Collect every key that appears anywhere in the diff timeline.
    affected = set()
    for nv in new_versions:
        for dr in nv.get("diff_rows", []):
            r = dr.get("curr") or dr.get("prev")
            if r is None:
                continue
            k = r.get(key)
            if k is not None and k != "":
                affected.add(k)

    # Baseline view for the first version: only rows whose key changed
    # somewhere downstream. Original row order is preserved.
    v0_rows = versions[0].get(data_key, []) if versions else []
    new_versions[0][data_key] = [r for r in v0_rows if r.get(key) in affected]

    return new_versions


class DiffTable:
    """Generate an interactive, version-diffing HTML page.

    Parameters
    ----------
    source : str, Path, or dict — either a path to a versioned JSON file
        (``{"versions": [...]}``) or a pre-built dict in the same shape.
        Use :meth:`from_files` to build the dict from N separate JSON files.
    title : str — page/header title.
    key : str — field used to match rows across versions. Defaults to the first field.
    output : str or Path or None — output HTML path. Defaults to ``<title>.html``
        next to the source file, or in the current directory if ``source`` is a dict.
    columns : list[str] or None — explicit columns to display. Auto-detected if None.
    hide_columns : list[str] or None — columns to omit from the rendered
        table (applied after auto-detection / explicit ``columns``). The
        key column is always shown regardless. Useful when the source data
        carries metadata fields you don't want in the diff view.
    note_field : str — field shown in the side panel. Default "note".
    changes_only : bool — if True, pre-compute per-version diffs in Python and
        embed only changed rows in the output (added / modified / removed
        relative to the previous version). The "Changes only" toggle is
        omitted from the UI. The first version renders empty since it has
        nothing to compare against. Trades baseline-view fidelity for a
        much smaller, faster-loading HTML file when most rows are unchanged
        across versions.
    """

    def __init__(self, source, *, title="Diff Explorer", key=None,
                 output=None, columns=None, hide_columns=None,
                 note_field="note", changes_only=True):
        if isinstance(source, dict):
            self.json_source = None
            self._data = source
        else:
            self.json_source = Path(source)
            self._data = None
        self.title = title
        self._key = key
        self._output = Path(output) if output else None
        self._columns = columns
        self._hide_columns = set(hide_columns) if hide_columns else set()
        self.note_field = note_field
        self.changes_only = changes_only

    @classmethod
    def from_files(cls, files, *, data_field=None, date_field=None,
                   row_filter=None, **kwargs):
        """Build a DiffTable from N separate JSON files, one version per file.

        Each ``files`` entry may be:
          * a path/str — version label defaults to the file stem.
          * ``(path, version)`` — explicit version label.
          * ``(path, version, date)`` — version label and date string.

        Each file's JSON may be either a list (used directly as the data
        array) or a dict whose first list-valued top-level key holds the data
        array. Pass ``data_field`` to pin that key explicitly.

        ``date_field`` names a top-level key in each source JSON whose value
        is used as the version date when the entry tuple doesn't supply one
        explicitly (e.g. ``date_field="timestamp"`` for files that embed
        their own provenance metadata). Without a date the timeline rail
        falls back to uniform spacing.

        ``row_filter`` is an optional callable applied to each row in every
        file; rows for which it returns falsy are dropped before the version
        entry is built. Useful for excluding categories of rows (e.g. high-
        speed signal sections) without rewriting the source JSON files.

        Returns a DiffTable bound to an in-memory ``{"versions": [...]}`` doc.
        """
        resolved = data_field
        versions = []
        for entry in files:
            if isinstance(entry, (str, Path)):
                path, version, date = Path(entry), Path(entry).stem, None
            else:
                items = list(entry)
                path = Path(items[0])
                version = items[1] if len(items) > 1 and items[1] else path.stem
                date = items[2] if len(items) > 2 else None

            with path.open("r", encoding="utf-8") as f:
                doc = json.load(f)

            if isinstance(doc, list):
                rows = doc
                if resolved is None:
                    resolved = "data"
            else:
                if resolved is None:
                    resolved = _detect_data_key(doc)
                rows = doc.get(resolved, [])
                if date is None and date_field:
                    date = doc.get(date_field)

            if row_filter is not None:
                rows = [r for r in rows if row_filter(r)]

            ver = {"version": version}
            if date:
                ver["date"] = date
            ver[resolved] = rows
            versions.append(ver)

        return cls({"versions": versions}, **kwargs)

    def generate(self):
        if self._data is not None:
            data = self._data
        else:
            with self.json_source.open("r", encoding="utf-8") as f:
                data = json.load(f)

        if "versions" not in data or not data["versions"]:
            raise ValueError("JSON must contain a non-empty 'versions' array.")

        versions = data["versions"]
        data_key = _detect_data_key(versions[0])
        all_cols = _detect_columns(versions, data_key)

        key = self._key or all_cols[0]
        if key not in all_cols:
            raise ValueError(f"Key field '{key}' not found in data.")

        if self._columns:
            display_cols = [c for c in self._columns if c in all_cols]
        else:
            display_cols = [c for c in all_cols if c != self.note_field]

        # Drop hidden columns. The key column is always shown — it's
        # re-inserted at index 0 below regardless.
        if self._hide_columns:
            display_cols = [c for c in display_cols if c not in self._hide_columns]

        if key in display_cols:
            display_cols.remove(key)
        display_cols.insert(0, key)

        if self._output:
            out_path = self._output
        elif self.json_source:
            out_path = self.json_source.parent / f"{self.title}.html"
        else:
            out_path = Path.cwd() / f"{self.title}.html"
        html = self._render(versions, data_key, key, display_cols)
        out_path.write_text(html, encoding="utf-8")
        return out_path

    def _render(self, versions, data_key, key, display_cols):
        # Skip diff-stripping when there's nothing to diff against:
        # _build_diff_only_versions filters the first version's data to
        # rows that change *somewhere downstream*, which is the empty
        # set with a single version — and an empty table is not what
        # the caller wants.
        if self.changes_only and len(versions) > 1:
            versions = _build_diff_only_versions(versions, data_key, key,
                                                 note_field=self.note_field)

        col_headers = "".join(
            f"<th>{self._col_label(c)}</th>" for c in display_cols
        )
        col_labels = [c.replace("_", " ").title() for c in display_cols]
        script = JS_TEMPLATE.substitute(
            KEY=json.dumps(key),
            DATA_KEY=json.dumps(data_key),
            COLS=json.dumps(display_cols),
            COL_LABELS=json.dumps(col_labels),
            NOTE_FIELD=json.dumps(self.note_field),
            VERSIONS=json.dumps(versions),
            CHANGES_ONLY=json.dumps(self.changes_only),
        )
        return HTML_TEMPLATE.substitute(
            TITLE=_esc(self.title),
            STYLE=STYLE,
            COL_HEADERS=col_headers,
            SCRIPT=script,
            TOGGLE_PILL="" if self.changes_only else TOGGLE_PILL_HTML,
        )

    @staticmethod
    def _col_label(name):
        return _esc(name.replace("_", " ").title())


class SpreadsheetConverter:
    """Convert xlsx/csv files to the versioned JSON format used by DiffTable.

    Parameters
    ----------
    source : str or Path — path to an .xlsx, .xls, or .csv file.
    output_dir : str or Path or None — defaults to the source file's directory.
    version : str — version label (e.g. "v1.0").
    date : str or None — optional date string.
    data_key : str — key name for the data array. Default "data".
    """

    def __init__(self, source, *, output_dir=None, version="v1.0", date=None,
                 data_key="data"):
        self.source = Path(source)
        self.output_dir = Path(output_dir) if output_dir else self.source.parent
        self.version = version
        self.date = date
        self.data_key = data_key

    def convert(self, append=False):
        """Convert the source file and write JSON files. Returns list[Path]."""
        suffix = self.source.suffix.lower()
        if suffix == ".csv":
            sheets = {"": self._read_csv()}
        elif suffix in (".xlsx", ".xls"):
            sheets = self._read_excel()
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

        stem = self.source.stem
        outputs = []
        for sheet_name, rows in sheets.items():
            filename = f"{stem}_{_safe_name(sheet_name)}.json" if sheet_name else f"{stem}.json"
            out_path = self.output_dir / filename

            version_entry = _make_version_entry(self.version, self.date, self.data_key, rows)

            if append and out_path.exists():
                with out_path.open("r", encoding="utf-8") as f:
                    existing = json.load(f)
                existing["versions"].append(version_entry)
            else:
                existing = {"versions": [version_entry]}

            with out_path.open("w", encoding="utf-8") as f:
                json.dump(existing, f, indent=2, ensure_ascii=False)

            outputs.append(out_path)

        return outputs

    def _read_csv(self):
        with self.source.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader]

    def _read_excel(self):
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(self.source, read_only=True, data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            rows = _read_sheet_rows(wb[name])
            if rows:
                sheets[name] = rows
        wb.close()
        return sheets


class ExcelDiff:
    """Scan a directory of xlsx files -> JSON -> HTML diffs.

    Files are sorted by name to determine version order. Sheets with the same
    name across files become successive versions.

    Parameters
    ----------
    directory : str or Path — directory containing .xlsx files.
    output_dir : str or Path or None — defaults to ``directory``.
    key : str or None — row identity field. Defaults to the first column.
    data_key : str — key for the data array. Default "data".
    note_field : str — field shown in the side panel. Default "note".
    version_func : callable or None — ``(filepath) -> (version_label, date | None)``.
    recent : int or None — only use the last N files (by name). None means all.
    changes_only : bool — forwarded to :class:`DiffTable`. See its docstring.
    """

    def __init__(self, directory, *, output_dir=None, key=None, data_key="data",
                 note_field="note", version_func=None, recent=None,
                 changes_only=True):
        self.directory = Path(directory)
        self.output_dir = Path(output_dir) if output_dir else self.directory
        self.key = key
        self.data_key = data_key
        self.note_field = note_field
        self.version_func = version_func or (lambda fp: (fp.stem, None))
        self.recent = recent
        self.changes_only = changes_only

    def run(self):
        """Execute the full pipeline: xlsx -> JSON -> HTML. Returns list[Path]."""
        openpyxl = _require_openpyxl()

        files = sorted(
            (f for f in self.directory.iterdir()
             if f.is_file() and f.suffix.lower() == ".xlsx"),
            key=lambda f: f.name,
        )
        if not files:
            raise FileNotFoundError(f"No .xlsx files found in {self.directory}")

        if self.recent is not None:
            files = files[-self.recent:]

        self.output_dir.mkdir(parents=True, exist_ok=True)

        # sheet_name -> {"versions": [...]}
        sheet_data = {}

        for filepath in files:
            version_label, date_str = self.version_func(filepath)
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                rows = _read_sheet_rows(wb[sheet_name])
                if not rows:
                    continue

                version_entry = _make_version_entry(
                    version_label, date_str, self.data_key, rows
                )

                safe = _safe_name(sheet_name)
                if safe not in sheet_data:
                    sheet_data[safe] = {"name": sheet_name, "versions": []}
                sheet_data[safe]["versions"].append(version_entry)

            wb.close()

        html_outputs = []
        for safe_name, info in sheet_data.items():
            json_path = self.output_dir / f"{safe_name}.json"
            with json_path.open("w", encoding="utf-8") as f:
                json.dump({"versions": info["versions"]}, f, indent=2,
                          ensure_ascii=False)

            dt = DiffTable(
                json_path,
                title=info["name"],
                key=self.key,
                note_field=self.note_field,
                changes_only=self.changes_only,
            )
            html_outputs.append(dt.generate())

        return html_outputs


class ZipDiff:
    """Build a branching git history from a folder of zip files.

    Creates a dedicated ``work_dir`` (default ``zip_dir/Diff``) and extracts
    every zip into it; the git repo lives inside ``work_dir``. The original
    zip files in ``zip_dir`` are never touched. ``work_dir`` is wiped and
    recreated on each run, so reruns are safe and idempotent.

    Workflow: the repo starts with an empty ``main``. For each zip (sorted
    by name), a branch is created; every top-level entry (subfolder or file)
    in the zip becomes its own commit on the branch (skipping entries that
    didn't change), then the branch is merged back into ``main`` with
    ``--no-ff``. The result: a clean per-zip merge history on ``main``, with
    per-subfolder granularity inside each branch.

    Parameters
    ----------
    zip_dir : str or Path — folder containing .zip files (read-only input).
    work_dir : str or Path — where files are unzipped and the git repo lives.
        If relative, resolved against ``zip_dir``. Default ``"var"``.
    message_func : callable or None — ``(filepath) -> merge commit body``.
        Defaults to the zip filename stem. The commit subject is always the
        branch name; ``message_func`` provides the body paragraph that follows.
    branch_func : callable or None — ``(filepath) -> branch name``.
        Default picks the first ``_``-token matching ``v\\d+(\\.\\d+)+`` and
        falls back to the full stem if no version token is found. When several
        zips yield the same raw name, every occurrence is patch-suffixed with
        ``.0``, ``.1``, … in sort order (e.g. two ``v1.2`` zips become
        ``v1.2.0`` and ``v1.2.1``).
    author_name : str or None — override git author name. Defaults to your
        global ``user.name`` from ``git config``.
    author_email : str or None — override git author email. Defaults to your
        global ``user.email`` from ``git config``.
    """

    def __init__(self, zip_dir, *, work_dir="var", message_func=None,
                 branch_func=None, author_name=None, author_email=None):
        self.zip_dir = Path(zip_dir)
        work = Path(work_dir)
        self.work_dir = work if work.is_absolute() else self.zip_dir / work
        self.message_func = message_func or (lambda fp: fp.stem)
        self.branch_func = branch_func or _default_branch_func
        self.author_name = author_name
        self.author_email = author_email

    def run(self):
        """Execute the pipeline. Returns the list of merge commit messages in order."""
        zips = sorted(
            (f for f in self.zip_dir.iterdir()
             if f.is_file() and f.suffix.lower() == ".zip"),
            key=lambda f: f.name,
        )
        if not zips:
            raise FileNotFoundError(f"No .zip files found in {self.zip_dir}")

        if self.work_dir.exists():
            shutil.rmtree(self.work_dir)
        self.work_dir.mkdir(parents=True)

        self._git("init", "-b", "main")
        if self.author_name:
            self._git("config", "user.name", self.author_name)
        if self.author_email:
            self._git("config", "user.email", self.author_email)
        if _GITIGNORE_SRC.exists():
            shutil.copy(_GITIGNORE_SRC, self.work_dir / ".gitignore")
            self._git("add", ".gitignore")
        self._git("commit", "--allow-empty", "-m", "Initial commit")

        branches = self._resolve_branches(zips)

        merges = []
        for zip_path, branch in zip(zips, branches):
            self._git("checkout", "-b", branch)

            with zipfile.ZipFile(zip_path) as zf:
                visible = [n for n in zf.namelist() if _is_visible(n)]
                wrapper = _detect_wrapper(visible)
                zip_entries = _top_level_entries(visible, wrapper)
                current_entries = {
                    p.name for p in self.work_dir.iterdir()
                    if p.name not in (".git", ".gitignore")
                }
                all_entries = sorted(current_entries | zip_entries)

                had_commit = False
                for entry in all_entries:
                    self._replace_entry(zf, entry, visible, zip_entries, wrapper)
                    self._git("add", "-A", "--", entry)
                    if self._has_staged_changes():
                        self._git("commit", "-m", entry)
                        had_commit = True

            if not had_commit:
                self._git("commit", "--allow-empty", "-m", "(no changes)")

            self._git("checkout", "main")
            body = self.message_func(zip_path)
            self._git("merge", "--no-ff", branch, "-m", branch, "-m", body)
            merges.append(f"{branch}\n\n{body}")

        return merges

    def _replace_entry(self, zf, entry, visible, zip_entries, wrapper):
        """Wipe a top-level entry from the work tree and re-extract it from the zip.

        If the zip has a single top-level wrapper folder, its prefix is stripped
        so the wrapper's children land directly in ``work_dir``.
        """
        entry_path = self.work_dir / entry
        if entry_path.exists():
            if entry_path.is_dir():
                shutil.rmtree(entry_path)
            else:
                entry_path.unlink()
        if entry not in zip_entries:
            return

        prefix = f"{wrapper}/{entry}" if wrapper else entry
        members = [n for n in visible
                   if n == prefix or n == prefix + "/"
                   or n.startswith(prefix + "/")]
        if wrapper:
            self._extract_stripped(zf, members, wrapper + "/")
        else:
            zf.extractall(self.work_dir, members=members)

    def _extract_stripped(self, zf, members, strip_prefix):
        """Extract ``members`` from ``zf`` into ``work_dir``, stripping ``strip_prefix``."""
        for m in members:
            if not m.startswith(strip_prefix):
                continue
            rel = m[len(strip_prefix):]
            if not rel:
                continue
            target = self.work_dir / rel
            if m.endswith("/"):
                target.mkdir(parents=True, exist_ok=True)
            else:
                target.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(m) as src, open(target, "wb") as dst:
                    shutil.copyfileobj(src, dst)

    def _resolve_branches(self, zips):
        """Compute branch names for ``zips``, patch-suffixing duplicates.

        Raw names shared by multiple zips become ``{name}.0``, ``{name}.1``, …
        in zip sort order. Unique names are returned as-is.
        """
        raw = [self.branch_func(zp) for zp in zips]
        counts = Counter(raw)
        next_idx = {}
        out = []
        for name in raw:
            if counts[name] > 1:
                i = next_idx.get(name, 0)
                next_idx[name] = i + 1
                out.append(f"{name}.{i}")
            else:
                out.append(name)
        return out

    def _has_staged_changes(self):
        """True if the index has any staged changes relative to HEAD."""
        result = subprocess.run(
            ["git", "diff", "--cached", "--quiet"],
            cwd=self.work_dir,
        )
        return result.returncode != 0

    def _git(self, *args):
        subprocess.run(
            ["git", *args],
            cwd=self.work_dir,
            check=True,
            capture_output=True,
            text=True,
        )


